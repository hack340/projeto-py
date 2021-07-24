# -*- coding: utf-8 -*-
"""

Este módulo contem funções para extrair dados das empresas a partir de planilhas em excell.

"""

#imports diversos
import openpyxl
from datetime import date

#Função principal, chamando funções especificas de linha e coluna
def get_raw_data_from_excel(folder):
    dic_year = {}
    
    #leitura do arquivo
    read = openpyxl.load_workbook(folder)
    sheet = read['BS P&L']
    
    #dados desejaveis começando pelas linhas e colunas setadas
    lin = 25
    col = 4
    
    dic_year = get_columns_data(sheet, lin, col)
    print(dic_year)
    return(dic_year)


#Função para varrer colunas e chamar função que varre as linhas
def get_columns_data(sheet, lin, col):
    dic_excel = {}

    #varre colunas dos anos, ano por ano da planilha
    while(col < 8):
        dic_excel[sheet.cell(15,col).value] = {sheet.cell(15,col).value: get_rows_data(sheet, col)}
        col += 1

    return(dic_excel)
    

#Função que varre as linhas para obter dados
def get_rows_data(sheet, col):
    dic_value = {}
    
    #Usado HardCode devido keys iguais usando dicionario
    dic_value['CAIXA'] = sheet.cell(25,col).value
    dic_value['APLICAÇÕES FINANCEIRAS'] = sheet.cell(26,col).value
    dic_value['CONTAS A RECEBER'] = sheet.cell(27,col).value
    dic_value['PDD'] = sheet.cell(28,col).value
    dic_value['ESTOQUE'] = sheet.cell(29,col).value
    dic_value['AC PARTES RELACIONADAS'] = sheet.cell(30,col).value
    dic_value['IR E CS DIFERIDOS'] = sheet.cell(31,col).value
    dic_value['CREDITOS FISCAIS A RECUPERAR'] = sheet.cell(32,col).value
    dic_value['ATIVOS DERIVATIVOS'] = sheet.cell(33,col).value
    dic_value['ADIANTAMENTOS'] = sheet.cell(34,col).value
    dic_value['OUTROS ATIVOS CIRCULANTES'] = sheet.cell(35,col).value
    dic_value['ANC PARTES RELACIONADAS'] = sheet.cell(37,col).value
    #invertido IR e CS devido conflito key dict
    dic_value['CS E IR DIFERIDOS'] = sheet.cell(38,col).value
    dic_value['CLIENTES'] = sheet.cell(39,col).value
    dic_value['REPACTO RISCO HIDROLÓGICO'] = sheet.cell(40,col).value
    dic_value['OUTROS ATIVOS NÃO CIRCULANTES'] = sheet.cell(41,col).value
    dic_value['IMOBILIZADO'] = sheet.cell(43,col).value
    dic_value['INVESTIMENTOS'] = sheet.cell(44,col).value
    dic_value['INTANGIVEIS'] = sheet.cell(45,col).value
    dic_value['EMPRÉSTIMOS E FINANCIAMENTO'] = sheet.cell(48,col).value
    dic_value['DEBENTURES'] = sheet.cell(49,col).value
    dic_value['PARTES RELACIONADAS'] = sheet.cell(51,col).value
    dic_value['EMPRÉSTIMO (CIRCULANTE DO LONGO PRAZO)'] = sheet.cell(52,col).value
    dic_value['FORNECEDORES'] = sheet.cell(53,col).value
    dic_value['SALARIOS E ENCARGOS SOCIAIS'] = sheet.cell(54,col).value
    dic_value['OBRIGAÇÕES TRIBUTÁRIAS'] = sheet.cell(55,col).value
    #Usado . final IR devido conflito key dict
    dic_value['IR. E CS DIFERIDOS'] = sheet.cell(56,col).value
    dic_value['DIVIDENDOS'] = sheet.cell(57,col).value
    dic_value['DERIVATIVOS'] = sheet.cell(58,col).value
    dic_value['ADIANTAMENTO DE CLIENTES'] = sheet.cell(59,col).value
    dic_value['PROVISÃO'] = sheet.cell(60,col).value
    dic_value['OUTROS'] = sheet.cell(61,col).value
    #financiamentos no plural para evitar key igual
    dic_value['EMPRÉSTIMOS E FINANCIAMENTOS'] = sheet.cell(63,col).value
    dic_value['DEBENTURES '] = sheet.cell(64,col).value
    dic_value['ADIANTAMENTO'] = sheet.cell(65,col).value
    dic_value['PROVISÃO '] = sheet.cell(66,col).value
    dic_value['CONCESSÕES A PAGAR'] = sheet.cell(67,col).value
    #Já tem outros, incluido espaço final devido conflito key dict
    dic_value['OUTROS '] = sheet.cell(68,col).value
    dic_value['PARTICIPAÇÃO MINORITÁRIA'] = sheet.cell(72,col).value
    dic_value['CAPITAL SOCIAL'] = sheet.cell(73,col).value
    dic_value['RESERVAS'] = sheet.cell(74,col).value
    dic_value['PREJUÍZO/LUCROS ACUMULADOS'] = sheet.cell(75,col).value
    dic_value['RECEITA BRUTA'] = sheet.cell(82,col).value
    dic_value['RECEITA LÍQUIDA'] = sheet.cell(83,col).value
    dic_value['CPV'] = sheet.cell(84,col).value
    dic_value['DEPRECIAÇÃO'] = sheet.cell(85,col).value
    dic_value['DESPESAS OPERACIONAIS'] = sheet.cell(88,col).value
    dic_value['DESPESAS ADMINISTRATIVAS'] = sheet.cell(89,col).value
    dic_value['DESPESAS PESSOAIS'] = sheet.cell(90,col).value
    dic_value['DESPESAS COMERCIAIS'] = sheet.cell(91,col).value
    dic_value['OUTRAS RECEITAS / DESPESAS'] = sheet.cell(92,col).value
    dic_value['DESPESA FINANCEIRA'] = sheet.cell(95,col).value
    dic_value['RECEITA FINANCEIRA'] = sheet.cell(96,col).value
    dic_value['OUTRAS RECEITAS / DESPESAS'] = sheet.cell(97,col).value
    dic_value['IMPOSTOS DIFERIDOS'] = sheet.cell(98,col).value
    dic_value['RESULTADO NÃO OPERACIONAL'] = sheet.cell(101,col).value
    dic_value['EQUIVALENCIA PATRIMONIAL'] = sheet.cell(102,col).value
    dic_value['IRPJ'] = sheet.cell(104,col).value
    dic_value['CSLL'] = sheet.cell(105,col).value
    #Já tem part. min., usado espaço final devido conflito key dict
    dic_value['PARTICIPAÇÃO MINORITÁRIA '] = sheet.cell(108,col).value
    
    return(dic_value)
    
    
#chamando função principal com endereço do arquivo
get_raw_data_from_excel("C:\Projeto\Template_Scorecard_Locked_2018.xlsm")

